"""Exporta a Excel los casos de sobre-emision detectados.

Uso:
    python3 tools/ticket_audit/export_excel.py [--days 90] [--out ruta.xlsx]
                                               [--cart-status APPROVED]
                                               [--include-under]

Por defecto: ultimos 90 dias, solo carritos APPROVED, solo OVER_ISSUED.
SOLO LECTURA sobre la base.
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from tools.ticket_audit.detector import (  # noqa: E402
    OVER_ISSUED, OVER_ISSUED_CORREGIDO, UNDER_ISSUED, TicketAuditor,
)

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
CRIT_FILL = PatternFill("solid", fgColor="F8CBAD")
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
BODY = Font(name=FONT, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '"$"#,##0;("$"#,##0);-'

COLUMNS = [
    ("Payment reference", "paymentReference", 20),
    ("Merchant", "merchantRef", 12),
    ("Nombre merchant", "merchantName", 26),
    ("Evento", "eventName", 40),
    ("Fecha compra", "cartDate", 18),
    ("Esperados", "expectedTickets", 11),
    ("Emitidos", "actualTickets", 10),
    ("Exceso", "delta", 9),
    ("Tickets totales", "totalTickets", 14),
    ("Monto carrito", "amountPaid", 15),
    ("Valor facial expuesto", "exposedAmount", 20),
    ("Metodo de pago", "paymentMethod", 20),
    ("Intentos de pago aprobados", "approvedPaymentIntents", 24),
    ("Primer ticket (UTC)", "createdAtFirst", 19),
    ("Ultimo ticket (UTC)", "createdAtLast", 19),
    ("Ventana (min)", "_spread_min", 13),
    ("Firma", "_firma", 34),
    ("Severidad", "severity", 11),
    ("Cart ID", "cartId", 26),
]


def firma(res) -> str:
    """Descripcion legible de por que sabemos que es duplicado."""
    clones = [g for g in res.duplicateGroups if g["type"] == "cloned_ticket_reference"]
    ident = [g for g in res.duplicateGroups if g["type"] == "same_act_and_id"]
    parts = []
    if clones:
        n = max(g["count"] for g in clones)
        parts.append(f"Ticket clonado x{n} (misma reference)")
    if ident:
        n = max(g["count"] for g in ident)
        parts.append(f"Mismo act+cedula x{n}")
    return " | ".join(parts) or "Conteo excede lo comprado"


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER


def build(results, days, cart_status, out_path, corregidos=()):
    wb = Workbook()

    # ---------- hoja detalle ----------
    ws = wb.active
    ws.title = "Duplicados"
    ws.append([c[0] for c in COLUMNS])
    for res in results:
        row = []
        for _, key, _w in COLUMNS:
            if key == "_spread_min":
                row.append(round(res.createdAtSpread / 60.0, 1))
            elif key == "_firma":
                row.append(firma(res))
            else:
                val = getattr(res, key, None)
                if isinstance(val, datetime):
                    val = val.replace(tzinfo=None)
                row.append(val)
        ws.append(row)

    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    for i, (_h, key, width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = width
        for cell in ws[letter][1:]:
            cell.font = BODY
            cell.border = BORDER
            if key in ("amountPaid", "exposedAmount"):
                cell.number_format = MONEY
            elif key in ("createdAtFirst", "createdAtLast", "cartDate"):
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif key == "delta" and isinstance(cell.value, int) and cell.value >= 2:
                cell.fill = CRIT_FILL
                cell.font = Font(name=FONT, size=10, bold=True)
    ws.row_dimensions[1].height = 30
    last = ws.max_row

    # ---------- hoja resumen ----------
    rs = wb.create_sheet("Resumen", 0)
    rs["A1"] = "Auditoria Carts <-> Tickets - Sobre-emision"
    rs["A1"].font = TITLE_FONT
    meta = [
        ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Ventana analizada", f"Ultimos {days} dias"),
        ("Filtro de carrito", cart_status or "todos"),
        ("Alcance", "Solo compras web (excluye tickets BO-GENERATED del backoffice)"),
        ("Base", "blast-prod (conexion de solo lectura)"),
    ]
    r = 3
    for k, v in meta:
        rs.cell(r, 1, k).font = Font(name=FONT, bold=True, size=10)
        rs.cell(r, 2, v).font = BODY
        r += 1

    r += 1
    rs.cell(r, 1, "Indicador").font = Font(name=FONT, bold=True, size=11)
    rs.cell(r, 2, "Valor").font = Font(name=FONT, bold=True, size=11)
    style_header(rs, r)
    head = r
    metrics = [
        ("Casos con sobre-emision", f"=COUNT(Duplicados!F2:F{last})", "#,##0"),
        ("Tickets en exceso", f"=SUM(Duplicados!H2:H{last})", "#,##0"),
        ("Valor facial expuesto", f"=SUM(Duplicados!K2:K{last})", MONEY),
        ("Monto de los carritos afectados", f"=SUM(Duplicados!J2:J{last})", MONEY),
        ("Caso con mayor exceso", f"=MAX(Duplicados!H2:H{last})", "#,##0"),
        ("Merchants afectados", f"=SUMPRODUCT((COUNTIF(Duplicados!B2:B{last},"
                               f"Duplicados!B2:B{last})>0)/COUNTIF(Duplicados!B2:B{last},"
                               f"Duplicados!B2:B{last}))", "#,##0"),
    ]
    for i, (label, formula, fmt) in enumerate(metrics, start=1):
        rs.cell(head + i, 1, label).font = BODY
        c = rs.cell(head + i, 2, formula)
        c.font = BODY
        c.number_format = fmt
        rs.cell(head + i, 1).border = BORDER
        c.border = BORDER

    note = head + len(metrics) + 2
    rs.cell(note, 1, "Notas").font = Font(name=FONT, bold=True, size=11)
    notas = [
        "Esperados = suma de (cantidad del item x ticketGroupAmount del act). El ticketGroupAmount "
        "es el multiplicador de las localidades grupales (palcos, combos): un act que emite N tickets.",
        "Emitidos = tickets con status APPROVED o VALIDATED. Los anulados (CANCELLED) no cuentan "
        "contra el aforo porque la anulacion sobrescribe el ticket y lo deja en precio 0.",
        "'Valor facial expuesto' es boleteria que podria escanearse en puerta, no dinero perdido. "
        "Se prorratea el monto del carrito entre los tickets esperados.",
        "'Firma' indica el patron: 'Ticket clonado' significa que los tickets de mas comparten el "
        "mismo tickets.reference que el original, es decir son copias del documento, no emisiones nuevas.",
        "Fechas de ticket en UTC (derivadas del ObjectId). 'Fecha compra' viene de cart.dateCreation, "
        "que el backend guarda en hora local de Colombia.",
    ]
    for i, n in enumerate(notas, start=1):
        c = rs.cell(note + i, 1, "- " + n)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rs.merge_cells(start_row=note + i, start_column=1, end_row=note + i, end_column=6)
        rs.row_dimensions[note + i].height = 28
    rs.column_dimensions["A"].width = 38
    rs.column_dimensions["B"].width = 24
    for col in "CDEF":
        rs.column_dimensions[col].width = 18

    # ---------- agregados ----------
    def pivot(title, keyfn, label):
        agg = defaultdict(lambda: [0, 0, 0.0])
        for res in results:
            a = agg[keyfn(res)]
            a[0] += 1
            a[1] += res.delta
            a[2] += res.exposedAmount
        sh = wb.create_sheet(title)
        sh.append([label, "Casos", "Tickets en exceso", "Valor facial expuesto"])
        for k, v in sorted(agg.items(), key=lambda x: -x[1][1]):
            sh.append([k or "(sin dato)", v[0], v[1], round(v[2])])
        style_header(sh)
        sh.column_dimensions["A"].width = 46
        for col, w in (("B", 12), ("C", 20), ("D", 24)):
            sh.column_dimensions[col].width = w
        for row in sh.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY
                cell.border = BORDER
            row[3].number_format = MONEY
        sh.freeze_panes = "A2"
        return sh

    pivot("Por evento", lambda r: r.eventName or r.eventId, "Evento")
    pivot("Por merchant", lambda r: f"{r.merchantRef} - {r.merchantName}", "Merchant")

    # ---------- corregidos a mano ----------
    if corregidos:
        cg = wb.create_sheet("Corregidos a mano")
        cg.append(["Payment reference", "Merchant", "Nombre merchant", "Evento",
                   "Comprados", "Emitidos en total", "Anulados", "Emitidos de más",
                   "Vigentes hoy", "Fecha emisión (UTC)", "Monto carrito"])
        for r in sorted(corregidos, key=lambda x: -x.correctedExcess):
            cg.append([r.paymentReference, r.merchantRef, r.merchantName, r.eventName,
                       r.expectedTickets, r.totalTickets, r.cancelledTickets,
                       r.correctedExcess, r.actualTickets,
                       (r.createdAtFirst.replace(tzinfo=None) if r.createdAtFirst else None),
                       r.amountPaid])
        style_header(cg)
        cg.freeze_panes = "A2"
        for col, w in (("A", 20), ("B", 12), ("C", 26), ("D", 40), ("E", 12), ("F", 17),
                       ("G", 11), ("H", 16), ("I", 13), ("J", 19), ("K", 15)):
            cg.column_dimensions[col].width = w
        for row in cg.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY
                cell.border = BORDER
            row[9].number_format = "yyyy-mm-dd hh:mm"
            row[10].number_format = MONEY
        nota = cg.max_row + 2
        c = cg.cell(nota, 1,
                    "Estas referencias hoy cuadran: alguien ya anuló los tickets sobrantes desde el "
                    "backoffice. Se listan porque prueban que la doble emisión ocurrió — sin esta hoja, "
                    "una corrección manual borra la evidencia de que el bug sigue activo. "
                    "Solo se incluyen casos donde todos los tickets nacieron en el mismo instante "
                    "(doble ejecución); si el excedente se emitió horas después queda como Ambiguo, "
                    "porque por conteos es indistinguible de una anulación con reemisión legítima.")
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        cg.merge_cells(start_row=nota, start_column=1, end_row=nota, end_column=8)
        cg.row_dimensions[nota].height = 74

    # ---------- cronologia ----------
    ch = wb.create_sheet("Cronologia")
    ch.append(["Dia (UTC)", "Casos", "Tickets en exceso"])
    byday = defaultdict(lambda: [0, 0])
    for res in results:
        d = (res.createdAtLast or res.createdAtFirst)
        if d:
            a = byday[d.strftime("%Y-%m-%d")]
            a[0] += 1
            a[1] += res.delta
    for d in sorted(byday):
        ch.append([d, byday[d][0], byday[d][1]])
    style_header(ch)
    for col, w in (("A", 16), ("B", 12), ("C", 20)):
        ch.column_dimensions[col].width = w
    for row in ch.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
            cell.border = BORDER
    ch.freeze_panes = "A2"

    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=90)
    ap.add_argument("--cart-status", default="APPROVED")
    ap.add_argument("--include-under", action="store_true")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    since = datetime.now(timezone.utc) - timedelta(days=args.days)
    auditor = TicketAuditor()
    wanted = {OVER_ISSUED} | ({UNDER_ISSUED} if args.include_under else set())

    print(f"Auditando ultimos {args.days} dias (carritos {args.cart_status})...", flush=True)
    todos = [r for r in auditor.audit(since=since, cart_status=args.cart_status)
             if r.status in wanted or r.status == OVER_ISSUED_CORREGIDO]
    results = [r for r in todos if r.status in wanted]
    corregidos = [r for r in todos if r.status == OVER_ISSUED_CORREGIDO]
    print(f"  hallazgos: {len(results)} vivos + {len(corregidos)} ya corregidos "
          f"- enriqueciendo con evento y pago...", flush=True)
    for r in todos:
        auditor.enrich_payment(r)
    results.sort(key=lambda r: (-r.delta, r.merchantRef or ""))

    out = args.out or str(Path.home() / "Downloads" /
                          f"auditoria_tickets_duplicados_{datetime.now():%Y%m%d}.xlsx")
    build(results, args.days, args.cart_status, out, corregidos)
    print(f"OK -> {out}")
    print(f"   casos: {len(results)} | tickets en exceso: {sum(r.delta for r in results)}")
    print(f"   corregidos a mano: {len(corregidos)} | "
          f"tickets emitidos de mas y anulados: {sum(r.correctedExcess for r in corregidos)}")


if __name__ == "__main__":
    main()
