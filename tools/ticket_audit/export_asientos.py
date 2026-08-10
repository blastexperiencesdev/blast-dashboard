"""Asientos numerados vendidos más de una vez.

En silletería numerada cada asiento es un act propio ("Fila A - Balcón 1 -
Asiento 12", capacidad 1), así que un asiento vendido dos veces deja dos
boletas vivas apuntando al mismo asiento. Este reporte las cruza.

Distingue las dos fallas, que se arreglan distinto:

- **Configuración duplicada**: el mismo asiento está cargado varias veces en el
  evento, cada copia con su propio inventario. El mapa sí bloquea el que se
  vendió, pero queda otro registro libre. Se arregla depurando el mapa.
- **Bloqueo fallido**: el mismo act se vendió dos veces. Ahí falló la reserva.

    python3 tools/ticket_audit/export_asientos.py

SOLO LECTURA sobre la data de negocio.
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from tools.ticket_audit.detector import (  # noqa: E402
    ALIVE_TICKET_STATUS, TicketAuditor, norm_status,
)

#: Solo etiquetas que identifican el asiento por completo. Una etiqueta como
#: "asiento 4" sin fila se repite en cada fila del recinto: agruparla produce
#: cientos de falsos positivos.
ETIQUETA_COMPLETA = re.compile(r"\bfila\b.*\basiento\b", re.I)

FONT = "Arial"
HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, bold=True, size=14, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
ROJO = PatternFill("solid", fgColor="FCE4E4")
AMBAR = PatternFill("solid", fgColor="FFF2CC")
MONEY = '"$"#,##0;("$"#,##0);-'


def normaliza(label: str) -> str:
    return re.sub(r"\s+", " ", (label or "")).strip().lower()


def recolectar(auditor: TicketAuditor) -> list:
    db = auditor.db

    asientos = {}
    for x in db.acts.find({}, {"label": 1, "event": 1}):
        lab = re.sub(r"\s+", " ", (x.get("label") or "")).strip()
        if not ETIQUETA_COMPLETA.search(lab):
            continue
        e = x.get("event")
        asientos[str(x["_id"])] = (
            str(e.id) if e is not None and hasattr(e, "id") else None, lab)

    ids = list(asientos)
    vivas = defaultdict(list)
    for i in range(0, len(ids), 2000):
        for t in db.tickets.find(
                {"actId": {"$in": ids[i:i + 2000]}},
                {"actId": 1, "status": 1, "paymentReference": 1, "name": 1,
                 "identification": 1, "email": 1, "cellphone": 1, "reference": 1,
                 "reservationDate": 1, "merchantReference": 1, "price": 1}):
            if norm_status(t.get("status")) not in ALIVE_TICKET_STATUS:
                continue
            ev, lab = asientos[t["actId"]]
            vivas[(ev, normaliza(lab))].append(t)

    conflictos = []
    for (ev, lab), ts in vivas.items():
        refs = {t.get("paymentReference") for t in ts}
        docs = {(t.get("identification") or "").strip() for t in ts}
        # Dos boletas del mismo comprador son reemisión o recompra, no conflicto.
        if len(refs) < 2 or len(docs) < 2:
            continue
        ts.sort(key=lambda t: t["_id"])
        conflictos.append(dict(
            evento=ev, asiento=lab, tickets=ts,
            acts=len({t["actId"] for t in ts}),
            causa=("Configuración duplicada" if len({t["actId"] for t in ts}) > 1
                   else "Bloqueo fallido"),
        ))
    return conflictos


def estilo(ws, fila=1):
    for c in ws[fila]:
        if c.value is not None:
            c.fill, c.font = HDR, HDRF
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER


def build(conf: list, auditor: TicketAuditor, out: str) -> str:
    wb = Workbook()
    corte = f"{datetime.now():%d/%m/%Y %H:%M}"
    hoy = datetime.now()

    def fecha_evento(c):
        fs = [t.get("reservationDate") for t in c["tickets"] if t.get("reservationDate")]
        return min(fs) if fs else None

    # ---------------- Resumen por evento ----------------
    rs = wb.active
    rs.title = "Resumen por evento"
    rs["A1"] = "Asientos numerados vendidos más de una vez"
    rs["A1"].font = TITLE
    rs["A2"] = (f"CORTE: {corte} · dos boletas vivas sobre el mismo asiento, "
                f"a compradores distintos · lectura de producción")
    rs["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")

    por_ev = defaultdict(lambda: dict(asientos=0, boletas=0, compradores=set(),
                                      config=0, bloqueo=0, fecha=None, merch=set()))
    for c in conf:
        a = por_ev[c["evento"]]
        a["asientos"] += 1
        a["boletas"] += len(c["tickets"])
        a["compradores"] |= {t.get("paymentReference") for t in c["tickets"]}
        a["config" if c["acts"] > 1 else "bloqueo"] += 1
        a["merch"] |= {t.get("merchantReference") for t in c["tickets"]}
        f = fecha_evento(c)
        if f and (a["fecha"] is None or f < a["fecha"]):
            a["fecha"] = f

    cols = [("Evento", 44), ("Merchant", 11), ("Fecha del evento", 17), ("Estado", 12),
            ("Asientos en conflicto", 13), ("Boletas involucradas", 13),
            ("Compras afectadas", 13), ("Por configuración duplicada", 15),
            ("Por bloqueo fallido", 14)]
    for j, (h, _w) in enumerate(cols, start=1):
        rs.cell(4, j, h)
    estilo(rs, 4)

    r = 5
    for ev, a in sorted(por_ev.items(), key=lambda x: -x[1]["asientos"]):
        futuro = a["fecha"] and a["fecha"] > hoy
        rs.append([
            (auditor.event_name(ev) or "(sin nombre)")[:70],
            ", ".join(sorted(x for x in a["merch"] if x))[:20],
            a["fecha"], "POR REALIZARSE" if futuro else "ya pasó",
            a["asientos"], a["boletas"], len(a["compradores"]),
            a["config"], a["bloqueo"],
        ])
        if futuro:
            for c_ in rs[r]:
                c_.fill = ROJO
        r += 1
    ult = r - 1
    rs.cell(r, 1, "TOTAL")
    for i in range(5, 10):
        L = get_column_letter(i)
        rs.cell(r, i, f"=SUM({L}5:{L}{ult})")
    for i, (_h, w) in enumerate(cols, start=1):
        L = get_column_letter(i)
        rs.column_dimensions[L].width = w
        for c_ in rs[L][3:]:
            c_.font = BOLD if c_.row == r else BODY
            c_.border = BORDER
            if L == "C":
                c_.number_format = "yyyy-mm-dd hh:mm"
            elif i >= 5:
                c_.number_format = "#,##0"
            if c_.row == r:
                c_.fill = TOTAL_FILL
    rs.row_dimensions[4].height = 42
    rs.freeze_panes = "A5"

    n = r + 2
    rs.cell(n, 1, "Cómo leer esto").font = Font(name=FONT, bold=True, size=11)
    notas = [
        "Cada asiento numerado es un act propio con capacidad 1, así que dos boletas vivas sobre el "
        "mismo asiento significan que se vendió dos veces.",
        "Solo se cuentan asientos cuya etiqueta identifica fila y asiento. Una etiqueta como "
        "«asiento 4» sin fila se repite en cada fila del recinto y agruparla daría falsos positivos.",
        "Se excluyen los casos donde las dos boletas son del mismo comprador: eso es una reemisión "
        "o una recompra, no un conflicto.",
        "«Por configuración duplicada»: el mismo asiento está cargado varias veces en el evento, cada "
        "copia con inventario propio. El mapa bloqueó el que se vendió, pero quedó otro libre. "
        "Se arregla depurando el mapa del evento.",
        "«Por bloqueo fallido»: el mismo act se vendió dos veces. Ahí falló la reserva del asiento.",
        "Las filas en rojo son eventos que todavía no se realizan: son los que hay que resolver "
        "primero, porque esas personas van a llegar a la misma silla.",
    ]
    for i, t in enumerate(notas, start=1):
        c_ = rs.cell(n + i, 1, "• " + t)
        c_.font = Font(name=FONT, size=9)
        c_.alignment = Alignment(wrap_text=True, vertical="top")
        rs.merge_cells(start_row=n + i, start_column=1, end_row=n + i, end_column=9)
        rs.row_dimensions[n + i].height = 26

    # ---------------- Asientos ----------------
    ws = wb.create_sheet("Asientos en conflicto")
    cols2 = [("Evento", 40), ("Asiento", 46), ("Fecha del evento", 17), ("Boletas", 9),
             ("Compras", 9), ("Causa", 24), ("Referencias de pago", 46)]
    for j, (h, _w) in enumerate(cols2, start=1):
        ws.cell(1, j, h)
    estilo(ws)
    for c in sorted(conf, key=lambda x: (auditor.event_name(x["evento"]) or "", x["asiento"])):
        refs = sorted({t.get("paymentReference") for t in c["tickets"]})
        ws.append([
            (auditor.event_name(c["evento"]) or "")[:70], c["asiento"][:60],
            fecha_evento(c), len(c["tickets"]), len(refs), c["causa"], ", ".join(refs)[:200],
        ])
    for i, (_h, w) in enumerate(cols2, start=1):
        L = get_column_letter(i)
        ws.column_dimensions[L].width = w
        for c_ in ws[L][1:]:
            c_.font = BODY
            c_.border = BORDER
            if L == "C":
                c_.number_format = "yyyy-mm-dd hh:mm"
            if L == "F" and c_.value == "Bloqueo fallido":
                c_.fill = AMBAR
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols2))}{ws.max_row}"

    # ---------------- Boletas ----------------
    bs = wb.create_sheet("Boletas afectadas")
    cols3 = [("Evento", 34), ("Asiento", 40), ("Orden de compra", 13), ("Payment reference", 18),
             ("Serial boleta", 16), ("Comprador", 26), ("Documento", 15), ("Correo", 28),
             ("Celular", 14), ("Valor", 13), ("Causa", 22)]
    for j, (h, _w) in enumerate(cols3, start=1):
        bs.cell(1, j, h)
    estilo(bs)
    for c in sorted(conf, key=lambda x: (auditor.event_name(x["evento"]) or "", x["asiento"])):
        for orden, t in enumerate(c["tickets"], start=1):
            bs.append([
                (auditor.event_name(c["evento"]) or "")[:60], c["asiento"][:60],
                "1º (conserva)" if orden == 1 else f"{orden}º (reubicar)",
                t.get("paymentReference"), t.get("reference"), t.get("name"),
                t.get("identification"), t.get("email"), t.get("cellphone"),
                float(t.get("price") or 0), c["causa"],
            ])
    for i, (_h, w) in enumerate(cols3, start=1):
        L = get_column_letter(i)
        bs.column_dimensions[L].width = w
        for c_ in bs[L][1:]:
            c_.font = BODY
            c_.border = BORDER
            if L == "J":
                c_.number_format = MONEY
            if L == "C" and isinstance(c_.value, str) and "reubicar" in c_.value:
                c_.fill = ROJO
    bs.freeze_panes = "D2"
    bs.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}{bs.max_row}"

    wb.save(out)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out")
    args = ap.parse_args()
    auditor = TicketAuditor()
    print("Cruzando asientos numerados...", flush=True)
    conf = recolectar(auditor)
    out = args.out or str(Path.home() / "Downloads" /
                          f"asientos_duplicados_{datetime.now():%Y%m%d}.xlsx")
    build(conf, auditor, out)
    print(f"OK -> {out}")
    print(f"   asientos en conflicto: {len(conf)}")
    print(f"   por configuración duplicada: {sum(1 for c in conf if c['acts'] > 1)}")
    print(f"   por bloqueo fallido:         {sum(1 for c in conf if c['acts'] == 1)}")


if __name__ == "__main__":
    main()
