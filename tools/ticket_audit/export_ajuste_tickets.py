"""Tabla de ajuste: servicio e IVA que se cobró, boleta por boleta.

Muchos tickets quedaron grabados con ticketService y tax en cero aunque al
comprador sí se le cobraron. Este reporte dice, para cada boleta, qué valor
debería tener, tomando como fuente el ítem del carrito (que es lo que la
pasarela efectivamente cobró).

    python3 tools/ticket_audit/export_ajuste_tickets.py --merchant ES029

SOLO LECTURA. Este script no escribe nada en la base: genera el Excel para que
la corrección la haga un humano con su propio proceso.
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from tools.ticket_audit.detector import ALIVE_TICKET_STATUS, TicketAuditor, norm_status  # noqa: E402

FONT = "Arial"
HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, bold=True, size=14, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
FIX = PatternFill("solid", fgColor="FCE4E4")
OKF = PatternFill("solid", fgColor="E2EFDA")
MONEY = '"$"#,##0;("$"#,##0);-'


#: Por debajo de este monto la diferencia es ruido de redondeo del reparto
#: (uno a tres pesos por boleta), no un servicio sin grabar. Marcarlas para
#: ajuste ensucia el cargue y descuadra la suma.
UMBRAL_AJUSTE = 10


def repartir(total: float, n: int) -> list:
    """Reparte un monto entre n boletas sin perder pesos por redondeo: todas
    llevan el valor redondeado y la última absorbe la diferencia."""
    if n <= 0:
        return []
    base = round(total / n)
    valores = [base] * n
    valores[-1] = round(total - base * (n - 1))
    return valores


def recolectar(auditor: TicketAuditor, merchant: str) -> list:
    db = auditor.db
    acts = auditor._load_acts()

    # tickets vivos de venta web, agrupados por referencia
    por_ref = defaultdict(list)
    for t in db.tickets.find({"merchantReference": merchant},
                             {"paymentReference": 1, "status": 1, "eventId": 1, "cartId": 1,
                              "price": 1, "ticketService": 1, "tax": 1, "reference": 1,
                              "actId": 1, "typeEntrance": 1, "name": 1}):
        if str(t.get("cartId") or "").startswith("BO-GENERATED"):
            continue
        if norm_status(t.get("status")) not in ALIVE_TICKET_STATUS:
            continue
        por_ref[t.get("paymentReference")].append(t)

    filas = []
    for c in db.carts.find({"merchantRef": merchant, "status": "APPROVED"},
                           {"reference": 1, "details": 1, "dateCreation": 1, "total": 1}):
        ref = c.get("reference")
        tickets = por_ref.get(ref)
        if not tickets:
            continue
        # valores cobrados por act, según el carrito
        cobrado = {}
        for d in c.get("details") or []:
            for it in d.get("items") or []:
                act = it.get("act")
                aid = str(act.id) if act is not None and hasattr(act, "id") else None
                if not aid:
                    continue
                n = int(it.get("quantity") or 0) * acts.get(aid, 1)
                cobrado[aid] = dict(n=n,
                                    serv=float(it.get("ticketService") or 0),
                                    iva=float(it.get("tax") or 0),
                                    precio=float(it.get("subtotal") or 0))
        por_act = defaultdict(list)
        for t in tickets:
            por_act[t.get("actId")].append(t)

        for aid, ts in por_act.items():
            info = cobrado.get(aid)
            if not info:
                # el ticket apunta a un act que el carrito no compró
                for t in ts:
                    filas.append(fila(c, t, None, None, "Act no está en el carrito"))
                continue
            servs = repartir(info["serv"], len(ts))
            ivas = repartir(info["iva"], len(ts))
            for t, s, i in zip(ts, servs, ivas):
                nota = "" if info["n"] == len(ts) else \
                    f"El carrito compró {info['n']} boletas de este act y hay {len(ts)} emitidas"
                filas.append(fila(c, t, s, i, nota))
    return filas


def fila(cart, t, serv_ok, iva_ok, nota) -> dict:
    serv_act = float(t.get("ticketService") or 0)
    iva_act = float(t.get("tax") or 0)
    return dict(
        ticketId=str(t["_id"]),
        serial=t.get("reference"),
        paymentReference=cart.get("reference"),
        eventId=t.get("eventId"),
        actId=t.get("actId"),
        localidad=t.get("typeEntrance"),
        asistente=t.get("name"),
        fecha=cart.get("dateCreation"),
        precio=float(t.get("price") or 0),
        serv_actual=serv_act,
        iva_actual=iva_act,
        serv_ok=serv_ok,
        iva_ok=iva_ok,
        dif=None if serv_ok is None else round((serv_ok - serv_act) + (iva_ok - iva_act)),
        nota=nota,
    )


def build(filas, auditor, merchant, nombre, out, filtro=None):
    if filtro:
        agujas = [f.upper() for f in filtro]
        filas = [f for f in filas
                 if any(x in (auditor.event_name(f["eventId"]) or "").upper() for x in agujas)]
    ajustar = [f for f in filas
               if f["dif"] is not None and abs(f["dif"]) >= UMBRAL_AJUSTE]
    redondeo = [f for f in filas
                if f["dif"] is not None and 0 < abs(f["dif"]) < UMBRAL_AJUSTE]
    corte = f"{datetime.now():%d/%m/%Y %H:%M}"
    wb = Workbook()

    # ---------------- Resumen ----------------
    rs = wb.active
    rs.title = "Resumen"
    rs["A1"] = f"Ajuste de servicio e IVA en tickets — {nombre}"
    rs["A1"].font = TITLE
    rs["A2"] = f"CORTE: {corte} (hora Colombia) · lectura de producción, no se modificó nada"
    rs["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")

    por_ev = defaultdict(lambda: [0, 0, 0.0, 0.0])
    for f in filas:
        a = por_ev[auditor.event_name(f["eventId"]) or f["eventId"]]
        a[0] += 1
        if f["dif"] is not None and abs(f["dif"]) >= UMBRAL_AJUSTE:
            a[1] += 1
            a[2] += (f["serv_ok"] or 0) - f["serv_actual"]
            a[3] += (f["iva_ok"] or 0) - f["iva_actual"]

    for j, h in enumerate(["Evento", "Boletas", "Boletas a ajustar", "Servicio a sumar",
                           "IVA a sumar", "Total a sumar"], start=1):
        rs.cell(4, j, h)
    style(rs, 4)
    r = 5
    for ev, a in sorted(por_ev.items(), key=lambda x: -x[1][2]):
        rs.append([ev[:60], a[0], a[1], round(a[2]), round(a[3]), f"=D{r}+E{r}"])
        r += 1
    ult = r - 1
    rs.cell(r, 1, "TOTAL")
    for i in range(2, 7):
        L = get_column_letter(i)
        rs.cell(r, i, f"=SUM({L}5:{L}{ult})")
    for i, w in enumerate((52, 11, 15, 17, 15, 16), start=1):
        L = get_column_letter(i)
        rs.column_dimensions[L].width = w
        for cell in rs[L][3:]:
            cell.font = BOLD if cell.row == r else BODY
            cell.border = BORDER
            if i >= 4:
                cell.number_format = MONEY
            elif i >= 2:
                cell.number_format = "#,##0"
            if cell.row == r:
                cell.fill = TOTAL_FILL
    rs.row_dimensions[4].height = 30

    n = r + 2
    rs.cell(n, 1, "Cómo usar este archivo").font = Font(name=FONT, bold=True, size=11)
    notas = [
        "La hoja «Ajuste por boleta» trae una fila por ticket vigente de venta web, con el servicio "
        "y el IVA que tiene hoy y el que debería tener.",
        "El valor correcto sale del ítem del carrito, que es lo que la pasarela efectivamente cobró. "
        "Se reparte entre las boletas de ese ítem; la última absorbe el redondeo para que la suma "
        "cuadre exacta con el carrito.",
        "La hoja «Solo las que hay que ajustar» filtra únicamente las boletas con el servicio sin "
        "grabar. Es la que sirve para cargar.",
        "Las boletas marcadas «redondeo» difieren en uno a tres pesos por el reparto entre boletas "
        "del mismo ítem. No se ajustan: tocarlas descuadraría la suma contra el carrito.",
        "Las columnas clave para el cargue son: ticketId (el _id del documento en la colección "
        "tickets), ticketService correcto e IVA correcto.",
        "Antes de escribir en producción: probar sobre una copia, y guardar los valores actuales por "
        "si hay que revertir. Este ajuste cambia data de negocio, no un registro de auditoría.",
        "Ajustar estos campos NO cambia lo que se le cobró al comprador ni lo que entró a la "
        "pasarela: solo corrige el registro de la boleta para que cuadre con lo cobrado.",
    ]
    for i, t in enumerate(notas, start=1):
        c = rs.cell(n + i, 1, "• " + t)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rs.merge_cells(start_row=n + i, start_column=1, end_row=n + i, end_column=6)
        rs.row_dimensions[n + i].height = 26

    # ---------------- Detalle ----------------
    cols = [("ticketId", 26), ("Serial de la boleta", 19), ("Payment reference", 19),
            ("Evento", 34), ("Localidad", 24), ("Asistente", 24), ("Fecha de compra", 18),
            ("Precio", 13), ("Servicio actual", 14), ("IVA actual", 13),
            ("Servicio correcto", 15), ("IVA correcto", 14), ("Ajuste total", 13),
            ("¿Ajustar?", 11), ("Observación", 40)]

    def hoja(ws, datos):
        for j, (h, _w) in enumerate(cols, start=1):
            ws.cell(1, j, h)
        style(ws, 1)
        for f in datos:
            ws.append([
                f["ticketId"], f["serial"], f["paymentReference"],
                (auditor.event_name(f["eventId"]) or "")[:60], f["localidad"], f["asistente"],
                f["fecha"], f["precio"], f["serv_actual"], f["iva_actual"],
                f["serv_ok"], f["iva_ok"], f["dif"],
                ("SÍ" if f["dif"] is not None and abs(f["dif"]) >= UMBRAL_AJUSTE
                 else "redondeo" if f["dif"] else "—" if f["dif"] == 0 else "revisar"),
                f["nota"],
            ])
        for i, (_h, w) in enumerate(cols, start=1):
            L = get_column_letter(i)
            ws.column_dimensions[L].width = w
            for cell in ws[L][1:]:
                cell.font = BODY
                cell.border = BORDER
                if L in ("H", "I", "J", "K", "L", "M"):
                    cell.number_format = MONEY
                elif L == "G":
                    cell.number_format = "yyyy-mm-dd hh:mm"
                if L == "N":
                    cell.alignment = Alignment(horizontal="center")
                    if cell.value == "SÍ":
                        cell.fill = FIX
                        cell.font = BOLD
                    elif cell.value in ("—", "redondeo"):
                        cell.fill = OKF
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
        ws.row_dimensions[1].height = 30

    hoja(wb.create_sheet("Ajuste por boleta"), filas)
    hoja(wb.create_sheet("Solo las que hay que ajustar"), ajustar)

    validar(wb)
    wb.save(out)
    return out, len(filas), len(ajustar)


def style(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = HDR
            cell.font = HDRF
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER


def validar(wb) -> None:
    """Un texto que empieza con '=' se guarda como fórmula y rompe el archivo."""
    malas = []
    for ws in wb.worksheets:
        for fila_ in ws.iter_rows():
            for cell in fila_:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and not (
                        len(v) > 1 and (v[1].isalnum() or v[1] in "(+-'\"")):
                    malas.append(f"{ws.title}!{cell.coordinate}: {v!r}")
    if malas:
        raise ValueError("Texto que Excel leería como fórmula inválida:\n  " + "\n  ".join(malas))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--merchant", default="ES029")
    ap.add_argument("--evento", action="append")
    ap.add_argument("--out")
    args = ap.parse_args()

    auditor = TicketAuditor()
    nombre = (auditor._load_merchants().get(args.merchant) or {}).get("name") or args.merchant
    print(f"Leyendo {nombre} ({args.merchant})...", flush=True)
    filas = recolectar(auditor, args.merchant)
    out = args.out or str(Path.home() / "Downloads" /
                          f"ajuste_servicio_iva_{args.merchant}_{datetime.now():%Y%m%d_%H%M}.xlsx")
    out, n, k = build(filas, auditor, args.merchant, nombre, out, args.evento)
    print(f"OK -> {out}")
    print(f"   boletas analizadas: {n} | a ajustar: {k}")


if __name__ == "__main__":
    main()
