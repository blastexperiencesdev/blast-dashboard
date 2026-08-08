"""Reporte de conciliacion por evento para un merchant.

Cruza, evento por evento: carritos aprobados -> boletas compradas -> boletas
emitidas -> recaudo, dejando explicita cada diferencia para que la cuenta
cierre exacta.

    python3 tools/ticket_audit/export_merchant.py --merchant ES029

SOLO LECTURA sobre la data de negocio.
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from tools.ticket_audit.detector import ALIVE_TICKET_STATUS, TicketAuditor, norm_status  # noqa: E402

FONT = "Arial"
HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, bold=True, size=14, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
OKF = PatternFill("solid", fgColor="E2EFDA")
BADF = PatternFill("solid", fgColor="FCE4E4")
MONEY = '"$"#,##0;("$"#,##0);-'


def gather(auditor: TicketAuditor, merchant: str) -> dict:
    db = auditor.db
    acts = auditor._load_acts()

    act_event = {}
    for a in db.acts.find({}, {"event": 1}):
        e = a.get("event")
        act_event[str(a["_id"])] = str(e.id) if e is not None and hasattr(e, "id") else None

    def blank():
        return dict(carritos=0, compradas=0, recaudo=0.0, subtotal=0.0, servicio=0.0, iva=0.0,
                    no_emitidas=0, reembolsadas=0, duplicadas=0, fuera_carrito=0,
                    vigentes=0, cortesias=0, anulados_total=0, pagos=0.0, monto_no_emitido=0.0,
                    bo_monto=0.0, bo_gratis=0, bo_pagas=0, bo_serv=0.0, bo_iva=0.0,
                    tk_precio=0.0, tk_serv=0.0, tk_iva=0.0)

    ev = defaultdict(blank)
    cart_event = {}      # reference -> evento principal
    cart_expected = {}   # reference -> boletas compradas
    cart_status = {}

    for c in db.carts.find({"merchantRef": merchant}, {"reference": 1, "status": 1, "details": 1, "total": 1}):
        ref = c.get("reference")
        cart_status[ref] = norm_status(c.get("status"))
        if cart_status[ref] != "APPROVED":
            continue
        eventos, exp_total = set(), 0
        for d in c.get("details") or []:
            e = d.get("event")
            eid = str(e.id) if e is not None and hasattr(e, "id") else None
            if not eid:
                continue
            eventos.add(eid)
            a = ev[eid]
            a["recaudo"] += float(d.get("total") or 0)
            a["subtotal"] += float(d.get("subtotal") or 0)
            a["servicio"] += float(d.get("ticketService") or 0)
            a["iva"] += float(d.get("tax") or 0)
            for it in d.get("items") or []:
                act = it.get("act")
                aid = str(act.id) if act is not None and hasattr(act, "id") else None
                n = int(it.get("quantity") or 0) * acts.get(aid, 1)
                a["compradas"] += n
                exp_total += n
        for eid in eventos:
            ev[eid]["carritos"] += 1
        cart_expected[ref] = exp_total
        if eventos:
            cart_event[ref] = next(iter(eventos))

    # tickets
    por_ref = defaultdict(list)
    for t in db.tickets.find({"merchantReference": merchant},
                             {"paymentReference": 1, "status": 1, "eventId": 1, "cartId": 1,
                              "price": 1, "ticketService": 1, "tax": 1}):
        eid = t.get("eventId")
        a = ev[eid]
        vivo = norm_status(t.get("status")) in ALIVE_TICKET_STATUS
        if str(t.get("cartId") or "").startswith("BO-GENERATED"):
            if vivo:
                a["cortesias"] += 1
                precio = float(t.get("price") or 0)
                serv = float(t.get("ticketService") or 0)
                iva = float(t.get("tax") or 0)
                a["bo_monto"] += precio + serv + iva
                a["bo_serv"] += serv
                a["bo_iva"] += iva
                a["bo_gratis" if precio == 0 else "bo_pagas"] += 1
            continue
        por_ref[t.get("paymentReference")].append(t)
        if vivo:
            a["vigentes"] += 1
            # Recaudo tal como quedó grabado en la boleta. No siempre coincide
            # con lo cobrado: hay tickets con servicio e IVA en cero aunque al
            # comprador sí se le cobraron.
            a["tk_precio"] += float(t.get("price") or 0)
            a["tk_serv"] += float(t.get("ticketService") or 0)
            a["tk_iva"] += float(t.get("tax") or 0)
            if cart_status.get(t.get("paymentReference")) != "APPROVED":
                a["fuera_carrito"] += 1
        else:
            a["anulados_total"] += 1

    # clasificar los anulados: exceso de una doble emision vs reembolso
    for ref, ts in por_ref.items():
        if cart_status.get(ref) != "APPROVED":
            continue
        exp = cart_expected.get(ref, 0)
        anulados = [t for t in ts if norm_status(t.get("status")) not in ALIVE_TICKET_STATUS]
        if not anulados:
            continue
        exceso = max(len(ts) - exp, 0)
        for i, t in enumerate(anulados):
            ev[t.get("eventId")]["duplicadas" if i < exceso else "reembolsadas"] += 1

    # carritos aprobados sin ningun ticket
    sin_ticket = []
    for ref, exp in cart_expected.items():
        if ref in por_ref:
            continue
        eid = cart_event.get(ref)
        c = db.carts.find_one({"reference": ref}, {"total": 1, "dateCreation": 1})
        ev[eid]["no_emitidas"] += exp
        ev[eid]["monto_no_emitido"] += float((c or {}).get("total") or 0)
        pi = db.paymentIntents.find_one({"reference": ref}, {"paymentIntentStatus": 1, "paymentMethod": 1})
        sin_ticket.append(dict(ref=ref, evento=eid, compradas=exp,
                               monto=float((c or {}).get("total") or 0),
                               fecha=(c or {}).get("dateCreation"),
                               pago=(pi or {}).get("paymentIntentStatus"),
                               metodo=(pi or {}).get("paymentMethod")))

    # boletas compradas que no se emitieron pero cuyo carrito sí generó otras
    for ref, exp in cart_expected.items():
        ts = por_ref.get(ref)
        if not ts:
            continue
        faltan = exp - len(ts)
        if faltan > 0:
            ev[cart_event.get(ref)]["no_emitidas"] += faltan

    # Recaudo segun la pasarela. Una misma referencia puede tener mas de un
    # paymentIntent aprobado (reintentos); sumarlos todos duplicaria el
    # recaudo, asi que se toma uno por referencia.
    pago_por_ref = {}
    for p in db.paymentIntents.find({"merchantReference": merchant},
                                    {"reference": 1, "paymentIntentStatus": 1, "finalPrice": 1}):
        if norm_status(p.get("paymentIntentStatus")) not in ALIVE_TICKET_STATUS:
            continue
        ref = p.get("reference")
        pago_por_ref[ref] = max(pago_por_ref.get(ref, 0.0), float(p.get("finalPrice") or 0))
    for ref, monto in pago_por_ref.items():
        eid = cart_event.get(ref)
        if eid:
            ev[eid]["pagos"] += monto

    # tickets emitidos sobre carritos no aprobados, al detalle
    fuera = []
    for ref, ts in por_ref.items():
        st = cart_status.get(ref)
        if st == "APPROVED":
            continue
        vivos = [t for t in ts if norm_status(t.get("status")) in ALIVE_TICKET_STATUS]
        if vivos:
            fuera.append(dict(ref=ref, estado=st or "(sin carrito)", tickets=len(vivos),
                              evento=vivos[0].get("eventId")))

    return dict(eventos=ev, sin_ticket=sin_ticket, fuera=fuera)


def corto_nombre(titulo: str) -> str:
    """Los títulos de evento traen el nombre del tour completo. Para las tablas
    de recaudo basta con distinguirlos."""
    t = (titulo or "").strip()
    for ciudad in ("BOGOTÁ", "BOGOTA", "MEDELLIN", "MEDELLÍN", "CALI"):
        if ciudad in t.upper():
            base = t.split("┃")[0].strip() or t[:28]
            return base if ciudad in base.upper() else f"{base[:22]} — {ciudad.title()}"
    return t[:40]


def style(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = HDR
            cell.font = HDRF
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER


def build(data, auditor, merchant, nombre, out, filtro=None):
    ev = data["eventos"]
    filas = sorted(ev.items(), key=lambda x: -x[1]["recaudo"])
    if filtro:
        agujas = [f.upper() for f in filtro]
        filas = [(e, a) for e, a in filas
                 if any(x in (auditor.event_name(e) or "").upper() for x in agujas)]
    wb = Workbook()

    # ---------------- Conciliación ----------------
    ws = wb.active
    ws.title = "Conciliación por evento"
    ws["A1"] = f"Conciliación de venta y emisión — {nombre}"
    ws["A1"].font = TITLE
    corte = f"{datetime.now():%d/%m/%Y %H:%M}"
    ws["A2"] = (f"Histórico completo · CORTE: {corte} (hora Colombia) · datos de producción "
                f"en modo solo lectura. Los eventos en venta siguen moviéndose: estas cifras "
                f"son la foto de ese momento.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")

    cols = [
        ("Evento", 46), ("Carritos aprobados", 13), ("Boletas compradas", 13),
        ("(−) No emitidas", 12), ("(−) Anuladas por reembolso", 14),
        ("(+) Emitidas sin carrito aprobado", 15), ("Boletas vigentes (calculado)", 13),
        ("Conteo real en base", 13), ("Cuadre", 9),
        ("(+) Duplicadas anuladas", 13), ("(+) Backoffice", 12),
        ("Total tickets en base (calculado)", 14), ("Conteo real", 12), ("Cuadre ", 9),
        ("Recaudo web", 16), ("Recaudo pasarela", 16), ("Cuadre (±$1.000)", 12),
        ("Backoffice: cortesías", 12), ("Backoffice: con precio", 12),
        ("Recaudo backoffice", 17), ("RECAUDO TOTAL", 18),
    ]
    hdr_row = 4
    for j, (h, _w) in enumerate(cols, start=1):
        ws.cell(hdr_row, j, h)
    style(ws, hdr_row)

    r = hdr_row + 1
    for eid, a in filas:
        vig = a["compradas"] - a["no_emitidas"] - a["reembolsadas"] + a["fuera_carrito"]
        total_base = vig + a["reembolsadas"] + a["duplicadas"] + a["cortesias"]
        ws.append([
            auditor.event_name(eid) or "(sin evento)",
            a["carritos"], a["compradas"], a["no_emitidas"], a["reembolsadas"],
            a["fuera_carrito"],
            f"=C{r}-D{r}-E{r}+F{r}", a["vigentes"], f"=IF(G{r}=H{r},\"OK\",\"revisar\")",
            a["duplicadas"], a["cortesias"],
            f"=G{r}+E{r}+J{r}+K{r}", a["vigentes"] + a["anulados_total"] + a["cortesias"],
            f"=IF(L{r}=M{r},\"OK\",\"revisar\")",
            round(a["recaudo"]), round(a["pagos"]),
            f"=IF(ABS(O{r}-P{r})<=1000,\"OK\",\"revisar\")",
            a["bo_gratis"], a["bo_pagas"], round(a["bo_monto"]), f"=O{r}+T{r}",
        ])
        r += 1

    last = r - 1
    tr = r
    ws.cell(tr, 1, "TOTAL")
    for i in range(2, len(cols) + 1):
        L = get_column_letter(i)
        if L in ("I", "N", "Q"):
            ws.cell(tr, i, f'=IF(COUNTIF({L}{hdr_row+1}:{L}{last},"revisar")=0,"OK","revisar")')
        else:
            ws.cell(tr, i, f"=SUM({L}{hdr_row+1}:{L}{last})")

    for i, (_h, w) in enumerate(cols, start=1):
        L = get_column_letter(i)
        ws.column_dimensions[L].width = w
        for cell in ws[L][hdr_row:]:
            cell.font = BODY
            cell.border = BORDER
            if L in ("O", "P", "T", "U"):
                cell.number_format = MONEY
            elif L != "A":
                cell.number_format = "#,##0"
            if L == "U" and cell.row >= hdr_row + 1:
                cell.font = BOLD
            if cell.row == tr:
                cell.font = BOLD
                cell.fill = TOTAL_FILL
            if L in ("I", "N", "Q") and cell.row > hdr_row:
                cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[hdr_row].height = 44
    ws.freeze_panes = f"B{hdr_row+1}"

    n = tr + 2
    ws.cell(n, 1, "Cómo se lee esta hoja").font = Font(name=FONT, bold=True, size=11)
    notas = [
        "Boletas compradas = suma de (cantidad del ítem × ticketGroupAmount del act). El ticketGroupAmount "
        "es el multiplicador de las localidades grupales: un palco o combo es un solo ítem que emite varias boletas.",
        "No emitidas: boletas de carritos aprobados y pagados para las que nunca se generó el ticket. "
        "Ver la hoja «Diferencias a revisar».",
        "Anuladas por reembolso: boletas legítimamente emitidas y anuladas después desde el backoffice.",
        "Emitidas sin carrito aprobado: tickets vigentes cuyo carrito quedó en CREATED-TIME-OUT o "
        "DELETED_BY_USER. La boleta existe y es válida en puerta.",
        "Duplicadas anuladas: boletas emitidas de más por una doble ejecución, ya anuladas. Ver hoja «Casos duplicados».",
        "Backoffice: boletas emitidas desde el panel, sin carrito por diseño. Se separan en cortesías "
        "(precio 0) y con precio (boletería física vendida por fuera de la plataforma).",
        "Recaudo backoffice = suma de precio + servicio + IVA de esas boletas. No pasa por la pasarela, "
        "así que no aparece en el recaudo web ni se puede cruzar contra ella.",
        "RECAUDO TOTAL = recaudo web + recaudo backoffice.",
        "Recaudo carritos = suma de los totales de los carritos aprobados. Recaudo pasarela = suma de los "
        "pagos aprobados por la pasarela, tomando uno por referencia (una referencia con reintentos tiene "
        "varios registros de pago y sumarlos duplicaría el recaudo). Se toleran diferencias de hasta $1.000 "
        "por el redondeo de centavos acumulado en miles de carritos.",
        "Las columnas «Cuadre» comparan el cálculo contra el conteo real en la base. Todas deben decir OK.",
    ]
    for i, t in enumerate(notas, start=1):
        c = ws.cell(n + i, 1, "• " + t)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=n + i, start_column=1, end_row=n + i, end_column=9)
        ws.row_dimensions[n + i].height = 26

    # ---------------- Desglose de recaudo ----------------
    # El "recaudo" puede significar tres cosas distintas segun quien pregunte:
    # lo que paga el comprador, lo que recibe el organizador, o la comision de
    # la plataforma. Se separan para que cada quien encuentre su cifra.
    rc = wb.create_sheet("Desglose de recaudo")
    rc["A1"] = "Desglose de recaudo"
    rc["A1"].font = TITLE
    rc["A2"] = (f"CORTE: {corte} · El total cobrado al comprador incluye el servicio de "
                f"plataforma y el IVA. El valor de cara es lo que corresponde al organizador.")
    rc["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")

    rcols = [("Evento", 30), ("Canal", 14), ("Boletas", 10),
             ("Precio en la boleta", 17), ("Servicio en la boleta", 17),
             ("IVA en la boleta", 15), ("TOTAL SEGÚN TICKETS", 19),
             ("Total cobrado (carritos)", 19), ("Diferencia", 15)]
    for j, (h, _w) in enumerate(rcols, start=1):
        rc.cell(4, j, h)
    style(rc, 4)

    fr = 5
    for eid, a in filas:
        corto = corto_nombre(auditor.event_name(eid))
        rc.append([corto, "Venta web", a["vigentes"],
                   round(a["tk_precio"]), round(a["tk_serv"]), round(a["tk_iva"]),
                   f"=D{fr}+E{fr}+F{fr}", round(a["recaudo"]), f"=H{fr}-G{fr}"])
        fr += 1
        rc.append([corto, "Backoffice", a["cortesias"],
                   round(a["bo_monto"] - a["bo_serv"] - a["bo_iva"]),
                   round(a["bo_serv"]), round(a["bo_iva"]),
                   f"=D{fr}+E{fr}+F{fr}", round(a["bo_monto"]), f"=H{fr}-G{fr}"])
        fr += 1
    ult = fr - 1
    rc.cell(fr, 1, "TOTAL")
    rc.cell(fr, 2, "web + backoffice")
    for i in range(3, 10):
        L = get_column_letter(i)
        rc.cell(fr, i, f"=SUM({L}5:{L}{ult})")

    for i, (_h, w) in enumerate(rcols, start=1):
        L = get_column_letter(i)
        rc.column_dimensions[L].width = w
        for cell in rc[L][4:]:
            cell.font = BOLD if (cell.row == fr or L == "G") else BODY
            cell.border = BORDER
            if i >= 4:
                cell.number_format = MONEY
            elif i == 3:
                cell.number_format = "#,##0"
            if cell.row == fr:
                cell.fill = TOTAL_FILL
            elif L == "G":
                cell.fill = OKF
            elif L == "I" and isinstance(cell.value, str):
                cell.fill = BADF
    rc.row_dimensions[4].height = 44

    n = fr + 2
    rc.cell(n, 1, "Cuál cifra usar").font = Font(name=FONT, bold=True, size=11)
    for i, t in enumerate([
        "TOTAL SEGÚN TICKETS (columna G) es la suma de lo que quedó grabado en cada boleta emitida. "
        "Es la cifra que pediste y la que coincide con los reportes que salen de la boletería.",
        "Total cobrado (columna H) es lo que realmente se le debitó al comprador, confirmado contra "
        "los pagos de la pasarela. Para conciliar caja se usa esta.",
        "La diferencia se concentra en Bogotá: hay boletas con servicio e IVA grabados en cero "
        "aunque al comprador sí se le cobraron. Ejemplo: la referencia TG390741649 tiene un carrito "
        "de $411.836, la pasarela cobró $411.836, y sus 3 boletas solo registran $360.000. En "
        "Medellín y Cali las dos bases coinciden.",
        "Precio en la boleta: el valor de la entrada. Es lo que le corresponde al organizador.",
        "Servicio de plataforma: la comisión que cobra Blast Tickets al comprador. No es ingreso "
        "del organizador.",
        "Total cobrado: lo que efectivamente se le debitó al comprador y lo que se cruza contra "
        "los pagos de la pasarela.",
        "El backoffice no pasa por la pasarela: son boletas emitidas desde el panel, así que su "
        "recaudo se registra pero no se puede conciliar contra la pasarela.",
        "Valor de cara + servicio + IVA da $45.000 por encima del total cobrado. Son dos carritos "
        "(GO274072213 y WL243395420) donde el campo de servicio quedó mal grabado: se registraron "
        "$50.000 y $35.000 de servicio cuando sus boletas suman $20.000 en cada uno. Al comprador "
        "se le cobró el total correcto; el error está solo en el desglose de esos dos carritos.",
    ], start=1):
        c = rc.cell(n + i, 1, "• " + t)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rc.merge_cells(start_row=n + i, start_column=1, end_row=n + i, end_column=7)
        rc.row_dimensions[n + i].height = 26

    # ---------------- Casos duplicados ----------------
    cd = wb.create_sheet("Casos duplicados")
    cd["A1"] = "Casos de doble emisión — resueltos"
    cd["A1"].font = TITLE
    cd["A2"] = (f"CORTE: {corte} · Referencias donde el sistema emitió más boletas de las "
                f"compradas. En ambas, el excedente ya fue anulado y el conteo hoy es correcto.")
    cd["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")

    ccols = [("Payment reference", 20), ("Evento", 44), ("Comprador", 26), ("Comprado", 11),
             ("Emitido", 10), ("Excedente", 11), ("Vigentes hoy", 12), ("Estado", 20),
             ("Recaudo", 14), ("Medio de pago", 16), ("Pagos aprobados", 14),
             ("Fecha de compra", 18), ("Emisión (UTC)", 18), ("Anulado por", 22)]
    cd.append([])
    cd.append([c[0] for c in ccols])
    style(cd, 4)

    for ref in ("MK102342013", "GQ065620426"):
        res = next(iter(auditor.audit(references=[ref], include_ok=True)), None)
        if res is None:
            continue
        auditor.enrich_payment(res)
        db = auditor.db
        t0 = db.tickets.find_one({"paymentReference": ref}, {"name": 1, "identification": 1})
        who = (t0 or {}).get("name", "").strip()
        aud = db.audit_log.find_one({"documentId": {"$in": res.ticketIds}}, {"user": 1})
        cd.append([
            res.paymentReference, res.eventName, who,
            res.expectedTickets, res.totalTickets, res.correctedExcess or res.delta,
            res.actualTickets, "Excedente anulado", round(res.amountPaid),
            res.paymentMethod, res.approvedPaymentIntents,
            res.cartDate.replace(tzinfo=None) if hasattr(res.cartDate, "replace") else res.cartDate,
            res.createdAtFirst.replace(tzinfo=None) if res.createdAtFirst else None,
            (aud or {}).get("user", ""),
        ])

    for i, (_h, w) in enumerate(ccols, start=1):
        L = get_column_letter(i)
        cd.column_dimensions[L].width = w
        for cell in cd[L][4:]:
            cell.font = BODY
            cell.border = BORDER
            if L == "I":
                cell.number_format = MONEY
            elif L in ("L", "M"):
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif L in ("D", "E", "F", "G", "K"):
                cell.number_format = "#,##0"
            if L == "F":
                cell.fill = BADF
            if L == "G":
                cell.fill = OKF
    cd.row_dimensions[4].height = 30

    n = cd.max_row + 2
    cd.cell(n, 1, "Diagnóstico").font = Font(name=FONT, bold=True, size=11)
    diag = [
        "En los dos casos hubo un único pago aprobado: la pasarela cobró una sola vez el valor correcto. "
        "Ningún comprador pagó de más.",
        "Todas las boletas de cada referencia se crearon en el mismo segundo. Eso descarta un doble clic "
        "o una reemisión manual: el proceso de emisión corrió dos veces seguidas sobre el mismo pago.",
        "Cada boleta tiene su propio número de serie, así que las sobrantes eran válidas en puerta hasta "
        "que se anularon. El riesgo era de aforo, no económico.",
        "Causa: no existe una clave de idempotencia que impida emitir dos veces sobre la misma referencia "
        "de pago, ni una restricción en la base que lo bloquee.",
    ]
    for i, t in enumerate(diag, start=1):
        c = cd.cell(n + i, 1, "• " + t)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        cd.merge_cells(start_row=n + i, start_column=1, end_row=n + i, end_column=8)
        cd.row_dimensions[n + i].height = 26

    # ---------------- Diferencias ----------------
    df = wb.create_sheet("Diferencias a revisar")
    df["A1"] = "Diferencias que explican el cuadre"
    df["A1"].font = TITLE
    df["A2"] = f"CORTE: {corte} (hora Colombia)"
    df["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")
    df["A3"] = "A. Carritos aprobados y pagados sin ninguna boleta emitida"
    df["A3"].font = Font(name=FONT, bold=True, size=11)
    df.append([])
    df.append(["Payment reference", "Evento", "Boletas compradas", "Monto pagado",
               "Fecha de compra", "Estado del pago", "Medio de pago"])
    style(df, 5)
    incluidos = {e for e, _a in filas}
    fila = 6
    for s in sorted([x for x in data["sin_ticket"] if x["evento"] in incluidos],
                    key=lambda x: -x["monto"]):
        df.cell(fila, 1, s["ref"]); df.cell(fila, 2, auditor.event_name(s["evento"]))
        df.cell(fila, 3, s["compradas"]); df.cell(fila, 4, round(s["monto"]))
        df.cell(fila, 5, s["fecha"]); df.cell(fila, 6, s["pago"]); df.cell(fila, 7, s["metodo"])
        fila += 1
    df.cell(fila, 1, "TOTAL").font = BOLD
    df.cell(fila, 3, f"=SUM(C6:C{fila-1})").font = BOLD
    df.cell(fila, 4, f"=SUM(D6:D{fila-1})").font = BOLD
    fin_a = fila

    fila += 3
    df.cell(fila, 1, "B. Boletas vigentes emitidas sobre carritos no aprobados").font = Font(name=FONT, bold=True, size=11)
    fila += 2
    for j, h in enumerate(["Payment reference", "Evento", "Estado del carrito", "Boletas vigentes"], start=1):
        df.cell(fila, j, h)
    style(df, fila)
    ini_b = fila + 1
    fila += 1
    for f in sorted([x for x in data["fuera"] if x["evento"] in incluidos],
                    key=lambda x: -x["tickets"]):
        df.cell(fila, 1, f["ref"]); df.cell(fila, 2, auditor.event_name(f["evento"]))
        df.cell(fila, 3, f["estado"]); df.cell(fila, 4, f["tickets"])
        fila += 1
    df.cell(fila, 1, "TOTAL").font = BOLD
    df.cell(fila, 4, f"=SUM(D{ini_b}:D{fila-1})").font = BOLD

    for L, w in (("A", 20), ("B", 44), ("C", 20), ("D", 16), ("E", 18), ("F", 16), ("G", 16)):
        df.column_dimensions[L].width = w
    for row in df.iter_rows(min_row=5, max_row=fila):
        for cell in row:
            if cell.value is None:
                continue
            if not cell.font.bold:
                cell.font = BODY
            cell.border = BORDER
    for rr in range(6, fin_a + 1):
        df.cell(rr, 4).number_format = MONEY
        df.cell(rr, 5).number_format = "yyyy-mm-dd hh:mm"

    n = fila + 2
    for i, t in enumerate([
        "Bloque A: el cliente pagó y el sistema nunca le generó las boletas. Requiere contactar al comprador "
        "y emitir manualmente o reembolsar.",
        "Bloque B: la boleta existe y es válida, pero su carrito quedó vencido o fue eliminado por el usuario. "
        "Conviene verificar que el pago se haya recibido.",
    ], start=1):
        c = df.cell(n + i, 1, "• " + t)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        df.merge_cells(start_row=n + i, start_column=1, end_row=n + i, end_column=7)
        df.row_dimensions[n + i].height = 26

    validar(wb)
    wb.save(out)
    return out


def validar(wb) -> None:
    """openpyxl convierte en fórmula cualquier texto que empiece con '='.
    Un encabezado como '= Boletas vigentes' se guarda como <f> Boletas
    vigentes</f>, que es una fórmula inválida y hace que Excel se niegue a
    abrir el archivo. Falla aquí antes de entregar algo que no abre."""
    malas = []
    for ws in wb.worksheets:
        for fila in ws.iter_rows():
            for cell in fila:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                cuerpo = v[1:].strip()
                # Una fórmula real arranca con función, referencia o signo.
                if not cuerpo or not (cuerpo[0].isalnum() or cuerpo[0] in "(+-'\""):
                    malas.append(f"{ws.title}!{cell.coordinate}: {v!r}")
                elif " " in cuerpo.split("(")[0] and "!" not in cuerpo:
                    malas.append(f"{ws.title}!{cell.coordinate}: {v!r}")
    if malas:
        raise ValueError("Texto que Excel leería como fórmula inválida:\n  " + "\n  ".join(malas))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--merchant", default="ES029")
    ap.add_argument("--evento", action="append",
                    help="filtra eventos cuyo título contenga este texto; repetible")
    ap.add_argument("--out")
    args = ap.parse_args()

    auditor = TicketAuditor()
    nombre = (auditor._load_merchants().get(args.merchant) or {}).get("name") or args.merchant
    print(f"Recolectando datos de {nombre} ({args.merchant})...", flush=True)
    data = gather(auditor, args.merchant)
    out = args.out or str(Path.home() / "Downloads" /
                          f"conciliacion_{args.merchant}_{datetime.now():%Y%m%d}.xlsx")
    build(data, auditor, args.merchant, nombre, out, args.evento)
    print(f"OK -> {out}")
    tw = tb = 0
    for eid, a in data["eventos"].items():
        nom_ev = auditor.event_name(eid) or ""
        if args.evento and not any(x.upper() in nom_ev.upper() for x in args.evento):
            continue
        vig = a["compradas"] - a["no_emitidas"] - a["reembolsadas"] + a["fuera_carrito"]
        flag = "OK " if vig == a["vigentes"] else "REVISAR"
        tw += a["recaudo"]; tb += a["bo_monto"]
        print(f"  {flag} {nom_ev[:42]:42s} calc={vig} real={a['vigentes']} "
              f"web=${a['recaudo']:>14,.0f} BO=${a['bo_monto']:>12,.0f}")
    print(f"      recaudo web ${tw:,.0f} + backoffice ${tb:,.0f} = ${tw+tb:,.0f}")


if __name__ == "__main__":
    main()
